from operator import index

import openpyxl

from models.ObjectApp import ObjectApp
from utils import logger


class FileExcel:

    def __init__(self, file) -> None:
        """Clase que abstrae un documento excel
        Args:
            file(str): Nombre del archivo
        """
        self.error = None

        try:
            self.ws = openpyxl.load_workbook(file).active
        except (FileNotFoundError, PermissionError, openpyxl.utils.exceptions.InvalidFileException) as e:
            if isinstance(e, FileNotFoundError):
                self.error = f"No se encuentra el archivo {file}"
            elif isinstance(e, PermissionError):
                self.error = f"No se tiene permisos para acceder al siguiente archivo: {file}"
            elif isinstance(e, openpyxl.utils.exceptions.InvalidFileException):
                self.error = f"El archivo {file} no es un archivo válido de Excel"
            logger.error(self.error)
        except Exception as e:
            self.error = f"Ocurrió un error inesperado: {e}\nNo se puede abrir el archivo"
            logger.error(self.error)
        else:
            self.columns = dict(self.__get_headers())

    def __get_headers(self) -> list:
        """Obtiene las columnas del archivo
        Returns:
            Lista de las columnas del archivo
        """
        columns = [
            (self.ws.cell(row=1, column=i).value, i) for i in range(1, self.ws.max_column + 1)
        ]
        return columns


class ActiveExcelFile:
    def __init__(self, file_excel: FileExcel) -> None:
        self._file = file_excel
        self._columns = None
        self._column_number = None
        self._objects = list()

    @property
    def column_number(self):
        return self._column_number

    @column_number.setter
    def column_number(self, column_number):
        self._column_number = column_number

    @property
    def columns(self):
        return self._columns

    @columns.setter
    def columns(self, columns):
        self._columns = columns

    @property
    def objects(self):
        return self._objects

    @objects.setter
    def objects(self, object_app: ObjectApp):
        self._objects.append(object_app)

    def get_count_values(self):
        return len(self._objects[0].values)

    def get_count_columns(self):
        return len(self._columns)

    def _get_column_values(self, column: str) -> list:
        """Obtiene los valores de una columna específica
        Args:
            column: Nombre de columna del archivo
        Returns:
            Lista con los valores de una columna específica
        """
        values_column = [
            self._file.ws.cell(row=i, column=self._file.columns[column]).value for i in
            range(2, self._file.ws.max_row + 1)
        ]
        result = filter(None, values_column)
        return list(result)

    def get_columns_values(self) -> None:
        """Obtiene los valores de una lista de columnas
        Returns:
            Lista de objetos ObjectApp
        """
        result = list()
        index_cells_empty = 0
        for i, column in enumerate(self._columns):
            index_cells_empty = i if column == self.column_number else None
            value_column = self._get_column_values(column)
            result.append(
                ObjectApp(
                    header=column,
                    values=value_column,
                )
            )
        if self.column_number not in self._columns:
            index_cells_empty = len(self.columns)
            result.append(
                ObjectApp(
                    header=self.column_number,
                    values=self._get_column_values(self.column_number),
                    status=False
                )
            )

        self._objects = result
        self.delete_empty_objects(index_cells_empty)

    def delete_empty_objects(self, index_numbers: int):
        cells_to_delete = [
            i for i, value in enumerate(self.objects[index_numbers].values)
            if value in [None, 0, ""]
        ]
        for object_app in self.objects:
            for cell in cells_to_delete:
                object_app.values.remove(cell)
