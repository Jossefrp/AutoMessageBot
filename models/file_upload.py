import openpyxl


class FileExcel:

    def __init__(self, file) -> None:
        """Clase que abstrae un documento excel
        Args:
            file(str): Nombre del archivo
        """

        self.ws = openpyxl.load_workbook(file).active
        self.columns = dict(self.get_headers())

    def get_headers(self) -> list:
        """Obtiene las columnas del archivo
        Returns:
            Lista de las columnas del archivo
        """
        columns = [
            (self.ws.cell(row=1, column=i).value, i) for i in range(1, self.ws.max_column + 1)
        ]
        return columns

    def get_values(self, header: str) -> list:
        """Obtiene los valores de una columna específica
        Args:
            header: Nombre de columna del archivo
        Returns:
            Lista con los valores de una columna específica
        """
        values_columns = [
            self.ws.cell(row=i, column=self.columns[header]).value for i in range(2, self.ws.max_row + 1)
        ]
        return values_columns
